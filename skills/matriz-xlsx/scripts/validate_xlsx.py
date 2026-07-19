#!/usr/bin/env python3
"""Check formulas, cached error values, and external references in an XLSX."""
import argparse, json
from pathlib import Path
from openpyxl import load_workbook

ERRORS = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!", "#NULL!"}
def main():
    parser = argparse.ArgumentParser(); parser.add_argument("input", type=Path); parser.add_argument("--out", type=Path); args = parser.parse_args()
    formulas = load_workbook(args.input, read_only=True, data_only=False); values = load_workbook(args.input, read_only=True, data_only=True); issues = []
    for formula_sheet, value_sheet in zip(formulas.worksheets, values.worksheets):
        for formula_row, value_row in zip(formula_sheet.iter_rows(), value_sheet.iter_rows()):
            for formula, value in zip(formula_row, value_row):
                if isinstance(value.value, str) and value.value in ERRORS: issues.append({"sheet": formula_sheet.title, "cell": formula.coordinate, "issue": value.value})
                if isinstance(formula.value, str) and "[" in formula.value and "]" in formula.value: issues.append({"sheet": formula_sheet.title, "cell": formula.coordinate, "issue": "external workbook reference"})
    report = json.dumps({"valid": not issues, "issues": issues}, indent=2) + "\n"
    if args.out: args.out.write_text(report, encoding="utf-8")
    else: print(report, end="")
    if issues: raise SystemExit(1)

if __name__ == "__main__": main()

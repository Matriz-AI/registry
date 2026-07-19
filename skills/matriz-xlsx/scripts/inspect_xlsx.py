#!/usr/bin/env python3
"""Report workbook structure, formulas, and visibility without changing a workbook."""
import argparse, json
from pathlib import Path
from openpyxl import load_workbook

def main():
    parser = argparse.ArgumentParser(); parser.add_argument("input", type=Path); parser.add_argument("--out", type=Path); args = parser.parse_args()
    book = load_workbook(args.input, read_only=True, data_only=False); sheets = []
    for sheet in book.worksheets:
        formulas = sum(1 for row in sheet.iter_rows() for cell in row if isinstance(cell.value, str) and cell.value.startswith("="))
        sheets.append({"name": sheet.title, "state": sheet.sheet_state, "rows": sheet.max_row, "columns": sheet.max_column, "formulas": formulas})
    payload = json.dumps({"file": args.input.name, "sheets": sheets, "definedNames": len(book.defined_names)}, indent=2) + "\n"
    if args.out: args.out.write_text(payload, encoding="utf-8")
    else: print(payload, end="")

if __name__ == "__main__": main()
